#!/usr/bin/env python3
"""
自达号业绩看板 - 数据提取脚本
读取 6月总自达号业绩.xlsx，生成 zidahao_data.json

数据口径：
  - 花名册（昵称/抖音号/机构归属）来自 Sheet2
  - 所有数值指标（GMV/消耗/退款等）从「6月数据」日流水实时聚合
  - 子机构：花开/集米/太古/九三/直属/半兆/久酒 共 7 个
"""

import json
import os
import subprocess as sp
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import load_workbook

XLSX_PATH = '/Users/xiaocao/CC/自达号业绩看板/6月总自达号业绩.xlsx'
OUTPUT_PATH = '/Users/xiaocao/CC/自达号业绩看板/data/zidahao_data.json'
ZDH_SUB_AGENCIES = ['花开自达号', '集米自达号', '太古自达号', '九三自达号', '直属自达号', '半兆自达号', '久酒自达号']
TREND_DAYS = 30


def safe_float(v):
    try:
        return float(v) if v is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def run():
    wb = load_workbook(XLSX_PATH, data_only=True)

    # ═══════════════════════════════════════════
    # === 1. 花名册（Sheet2）===
    # ═══════════════════════════════════════════
    ws_roster = wb['Sheet2']
    roster = {}          # douyin_id → {name, agency}
    roster_order = []    # 保持原始顺序

    for r in range(2, ws_roster.max_row + 1):
        name = ws_roster.cell(r, 1).value
        douyin_id = ws_roster.cell(r, 2).value
        agency = ws_roster.cell(r, 3).value

        if not name or str(name).strip() == 'None' or str(name).strip() == '':
            continue
        if str(name).strip() == '汇总':
            continue

        did = str(douyin_id).strip() if douyin_id else ''
        if not did or did == 'None':
            continue

        roster[did] = {
            '主播昵称': str(name).strip(),
            '机构': str(agency).strip() if agency and str(agency).strip() != 'None' else '其他',
        }
        roster_order.append(did)

    print(f'  花名册: {len(roster)} 达人')

    # ═══════════════════════════════════════════
    # === 2. 日流水（6月数据）===
    # ═══════════════════════════════════════════
    ws_live = wb['6月数据']

    # 按抖音号 + 日期维度累加
    daily_gmv = defaultdict(lambda: defaultdict(float))
    daily_paid = defaultdict(lambda: defaultdict(float))
    daily_refund = defaultdict(lambda: defaultdict(float))
    daily_ad = defaultdict(lambda: defaultdict(float))
    daily_commission = defaultdict(lambda: defaultdict(float))
    daily_duration = defaultdict(lambda: defaultdict(float))
    live_name_map = {}  # 抖音号 → 昵称（日流水覆盖最全）

    all_dates_set = set()

    for r in range(2, ws_live.max_row + 1):
        douyin_id_raw = ws_live.cell(r, 3).value   # C列：主播抖音号
        nickname_raw = ws_live.cell(r, 2).value     # B列：主播昵称
        dt_val = ws_live.cell(r, 4).value            # D列：直播开始时间

        if not douyin_id_raw or not dt_val:
            continue

        douyin_id = str(douyin_id_raw).strip()
        date_key = str(dt_val)[:10].replace('/', '-')
        all_dates_set.add(date_key)

        # 昵称映射
        if nickname_raw:
            live_name_map[douyin_id] = str(nickname_raw).strip()

        gmv = safe_float(ws_live.cell(r, 26).value)          # Z: 直播间成交金额
        paid = safe_float(ws_live.cell(r, 27).value)          # AA: 支付金额
        refund = safe_float(ws_live.cell(r, 32).value)        # AF: 退款金额
        ad_bind = safe_float(ws_live.cell(r, 44).value)       # AR: 投放消耗(绑定)
        ad_beitou = safe_float(ws_live.cell(r, 45).value)     # AS: 投放消耗(被投)
        commission = safe_float(ws_live.cell(r, 34).value)    # AH: 预估佣金支出
        duration = safe_float(ws_live.cell(r, 6).value)       # F: 直播时长(分钟)

        # 消耗：被投优先，0 回退绑定
        ad_cost = ad_beitou if ad_beitou > 0 else ad_bind

        daily_gmv[douyin_id][date_key] += gmv
        daily_paid[douyin_id][date_key] += paid
        daily_refund[douyin_id][date_key] += refund
        daily_ad[douyin_id][date_key] += ad_cost if gmv > 0 else 0
        daily_commission[douyin_id][date_key] += commission
        daily_duration[douyin_id][date_key] += duration

    all_dates = sorted(all_dates_set)
    print(f'  日流水: {len(daily_gmv)} 个抖音号, {len(all_dates)} 天')

    # ═══════════════════════════════════════════
    # === 3. 确定当前月份 ===
    # ═══════════════════════════════════════════
    latest_date = all_dates[-1]
    current_month = latest_date[:7]
    current_month_dates = [d for d in all_dates if d.startswith(current_month)]
    print(f'  当前月份: {current_month} | 最新日期: {latest_date} | 当月天数: {len(current_month_dates)}')

    # ═══════════════════════════════════════════
    # === 4. 达人月度指标（当月聚合）===
    # ═══════════════════════════════════════════

    def sum_month(daily_dict, douyin_id):
        return round(sum(
            daily_dict.get(douyin_id, {}).get(d, 0)
            for d in current_month_dates
        ), 2)

    def count_active_days(daily_dict, douyin_id):
        return sum(1 for d in current_month_dates
                   if daily_dict.get(douyin_id, {}).get(d, 0) > 0)

    # 对花名册中每个达人计算月度指标
    zdh_anchor_monthly = {}
    for douyin_id in roster:
        gmv_val = sum_month(daily_gmv, douyin_id)
        paid_val = sum_month(daily_paid, douyin_id)
        refund_val = sum_month(daily_refund, douyin_id)
        ad_val = sum_month(daily_ad, douyin_id)
        commission_val = sum_month(daily_commission, douyin_id)
        duration_total = sum_month(daily_duration, douyin_id)
        active_days = count_active_days(daily_gmv, douyin_id)
        settle_val = round(paid_val - refund_val, 2)

        info = roster[douyin_id]
        zdh_anchor_monthly[douyin_id] = {
            '主播昵称': live_name_map.get(douyin_id) or info['主播昵称'],
            '主播抖音号': douyin_id,
            '机构': info['机构'],
            '开播天数': active_days,
            '日均开播时长（小时）': round(duration_total / 60 / active_days, 1) if active_days > 0 else 0,
            '直播GMV': gmv_val,
            '直播支付GMV': paid_val,
            '直播退款GMV': refund_val,
            '直播结算GMV': settle_val,
            '结算率': round(settle_val / gmv_val, 4) if gmv_val > 0 else 0,
            'ROI': round(gmv_val / ad_val, 2) if ad_val > 0 else 0,
            '佣金支出': commission_val,
            '投放消耗金额': ad_val,
        }

    # ═══════════════════════════════════════════
    # === 5. 汇总卡片 ===
    # ═══════════════════════════════════════════
    zdh_total_gmv = sum(a['直播GMV'] for a in zdh_anchor_monthly.values())
    zdh_total_paid = sum(sum_month(daily_paid, did) for did in roster)
    zdh_total_refund = sum(a['直播退款GMV'] for a in zdh_anchor_monthly.values())
    zdh_total_settle = round(zdh_total_paid - zdh_total_refund, 2)
    zdh_total_ad = sum(a['投放消耗金额'] for a in zdh_anchor_monthly.values())

    summary = {
        '直播GMV': round(zdh_total_gmv, 2),
        '直播支付GMV': round(zdh_total_paid, 2),
        '直播退款GMV': round(zdh_total_refund, 2),
        '直播结算GMV': zdh_total_settle,
        '结算率': round(zdh_total_settle / zdh_total_gmv, 4) if zdh_total_gmv > 0 else 0,
        '消耗金额': round(zdh_total_ad, 2),
        'ROI': round(zdh_total_gmv / zdh_total_ad, 2) if zdh_total_ad > 0 else 0,
    }

    # ═══════════════════════════════════════════
    # === 6. 子机构汇总 ===
    # ═══════════════════════════════════════════
    zdh_sub_data = defaultdict(lambda: {
        'anchors': set(),
        'gmv': 0.0, 'refund': 0.0, 'paid': 0.0, 'ad': 0.0,
    })
    for douyin_id, a in zdh_anchor_monthly.items():
        sub = a['机构']
        zdh_sub_data[sub]['anchors'].add(douyin_id)
        zdh_sub_data[sub]['gmv'] += a['直播GMV']
        zdh_sub_data[sub]['refund'] += a['直播退款GMV']
        zdh_sub_data[sub]['paid'] += sum_month(daily_paid, douyin_id)
        zdh_sub_data[sub]['ad'] += a['投放消耗金额']

    zdh_sub_agencies = []
    for sub in ZDH_SUB_AGENCIES:
        if sub not in zdh_sub_data:
            continue
        d = zdh_sub_data[sub]
        n = len(d['anchors'])
        settle = round(d['paid'] - d['refund'], 2)
        zdh_sub_agencies.append({
            '机构': sub,
            '机构达人数': n,
            '直播GMV': round(d['gmv'], 2),
            '人均直播GMV': round(d['gmv'] / n, 2) if n > 0 else 0,
            '直播退款GMV': round(d['refund'], 2),
            '直播结算GMV': settle,
            '投放消耗金额': round(d['ad'], 2),
            'ROI': round(d['gmv'] / d['ad'], 2) if d['ad'] > 0 else 0,
        })

    # ═══════════════════════════════════════════
    # === 7. 趋势数据（近 30 天）===
    # ═══════════════════════════════════════════
    trend_dates_full = all_dates[-TREND_DAYS:] if len(all_dates) >= TREND_DAYS else all_dates
    trend_dates = [f"{int(d[5:7])}/{int(d[8:10])}" for d in trend_dates_full]
    date_map = {trend_dates_full[i]: trend_dates[i] for i in range(len(trend_dates_full))}

    # 自达号每日聚合
    zdh_gmv_daily = defaultdict(float)
    zdh_paid_daily = defaultdict(float)
    zdh_refund_daily = defaultdict(float)
    zdh_daily_by_sub = defaultdict(lambda: defaultdict(float))

    for d in trend_dates_full:
        for douyin_id, by_date in daily_gmv.items():
            gmv = by_date.get(d, 0)
            paid = daily_paid[douyin_id].get(d, 0)
            refund = daily_refund[douyin_id].get(d, 0)

            if douyin_id in roster:
                zdh_gmv_daily[d] += gmv
                zdh_paid_daily[d] += paid
                zdh_refund_daily[d] += refund
                sub = roster[douyin_id]['机构']
                zdh_daily_by_sub[sub][d] += gmv

    def daily_list(daily_dict):
        """将 defaultdict 转为按 trend_dates_full 顺序的列表（单位：万）"""
        return [round(daily_dict.get(d, 0) / 10000, 2) for d in trend_dates_full]

    # ═══════════════════════════════════════════
    # === 8. 子机构每日 ROI ===
    # ═══════════════════════════════════════════
    zdh_anchor_ids_by_sub = defaultdict(list)
    for douyin_id, info in roster.items():
        zdh_anchor_ids_by_sub[info['机构']].append(douyin_id)

    zdh_daily_roi_by_sub = {}
    for sub in ZDH_SUB_AGENCIES:
        aids = zdh_anchor_ids_by_sub.get(sub, [])
        roi_vals = []
        for d in trend_dates_full:
            gmv_sum = sum(daily_gmv.get(aid, {}).get(d, 0) for aid in aids)
            ad_sum = sum(daily_ad.get(aid, {}).get(d, 0) for aid in aids)
            roi_vals.append(round(gmv_sum / ad_sum, 2) if ad_sum > 0 else None)
        zdh_daily_roi_by_sub[sub] = roi_vals

    # 整体 ROI
    all_zdh_ids = list(roster.keys())
    zdh_daily_roi_overall = []
    for d in trend_dates_full:
        gmv_sum = sum(daily_gmv.get(aid, {}).get(d, 0) for aid in all_zdh_ids)
        ad_sum = sum(daily_ad.get(aid, {}).get(d, 0) for aid in all_zdh_ids)
        zdh_daily_roi_overall.append(round(gmv_sum / ad_sum, 2) if ad_sum > 0 else None)

    # ═══════════════════════════════════════════
    # === 9. 下探数据 ===
    # ═══════════════════════════════════════════

    # 9a. 子机构 Top5 达人分天数据
    zdh_top5_by_sub = {}
    for sub in ZDH_SUB_AGENCIES:
        sub_anchors = [
            (did, zdh_anchor_monthly[did]['直播GMV'])
            for did in roster
            if zdh_anchor_monthly[did]['机构'] == sub
        ]
        sub_anchors.sort(key=lambda x: x[1], reverse=True)
        top5 = []
        for douyin_id, _ in sub_anchors[:5]:
            info = roster[douyin_id]
            name = live_name_map.get(douyin_id) or info['主播昵称']
            vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
            top5.append({'name': str(name), 'douyin_id': douyin_id, 'daily_paid': vals})
        if top5:
            zdh_top5_by_sub[sub] = top5

    # 9b. 自达号下探明细（含 agency 字段）
    zdh_anchor_detail = []
    for douyin_id in roster:
        info = roster[douyin_id]
        vals = [round(daily_paid.get(douyin_id, {}).get(d, 0) / 10000, 2) for d in trend_dates_full]
        zdh_anchor_detail.append({
            'name': str(live_name_map.get(douyin_id) or info['主播昵称'] or douyin_id),
            'douyin_id': douyin_id,
            'agency': str(info.get('机构', '')),
            'daily_paid': vals
        })
    zdh_anchor_detail.sort(key=lambda x: sum(x['daily_paid']), reverse=True)

    # 9c. anchor_daily_paid（按抖音号）
    anchor_daily_paid_out = {}
    for douyin_id in daily_paid:
        anchor_daily_paid_out[douyin_id] = {
            date_map[d]: round(daily_paid[douyin_id].get(d, 0) / 10000, 2)
            for d in trend_dates_full
        }

    # 9d. anchor_daily_roi
    anchor_daily_roi = {}
    for douyin_id in daily_gmv:
        anchor_daily_roi[douyin_id] = {}
        for d in trend_dates_full:
            g = daily_gmv[douyin_id].get(d, 0)
            a = daily_ad.get(douyin_id, {}).get(d, 0)
            anchor_daily_roi[douyin_id][date_map[d]] = round(g / a, 2) if a > 0 else 0

    # ═══════════════════════════════════════════
    # === 10. 构建最终 JSON ===
    # ═══════════════════════════════════════════

    zdh_anchors_sorted = sorted(zdh_anchor_monthly.values(), key=lambda x: x['直播GMV'], reverse=True)

    dashboard_data = {
        'summary': summary,
        'sub_agencies': zdh_sub_agencies,
        'anchors': zdh_anchors_sorted,
        'daily_gmv': daily_list(zdh_gmv_daily),
        'daily_paid': daily_list(zdh_paid_daily),
        'daily_refund': daily_list(zdh_refund_daily),
        'daily_by_sub': {
            sub: daily_list(zdh_daily_by_sub.get(sub, defaultdict(float)))
            for sub in ZDH_SUB_AGENCIES
        },
        'pie_data': [
            {
                'name': sub,
                'value': round(sum(
                    gmv for d, gmv in zdh_daily_by_sub.get(sub, defaultdict(float)).items()
                    if d.startswith(current_month)
                ) / 10000, 2)
            }
            for sub in ZDH_SUB_AGENCIES
            if sub in zdh_daily_by_sub
        ],
        'top5_by_sub': zdh_top5_by_sub,
        'daily_roi_by_sub': zdh_daily_roi_by_sub,
        'daily_roi_overall': zdh_daily_roi_overall,
        'anchor_detail': zdh_anchor_detail,
        'anchor_daily_paid': anchor_daily_paid_out,
        'anchor_daily_roi': anchor_daily_roi,
        'trend_dates': trend_dates,
        'current_month': current_month,
        'latest_date': latest_date,
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

    print(f'\n✓ 自达号数据已更新到 {OUTPUT_PATH}')
    print(f'  - 当前月份: {current_month}')
    print(f'  - 汇总 GMV: ¥{summary["直播GMV"]:,.2f}')
    print(f'  - 结算 GMV: ¥{summary["直播结算GMV"]:,.2f}')
    print(f'  - 消耗金额: ¥{summary["消耗金额"]:,.2f}')
    print(f'  - ROI: {summary["ROI"]}')
    print(f'  - 达人数量: {len(zdh_anchors_sorted)}')
    print(f'  - 子机构: {len(zdh_sub_agencies)} 个')
    for sa in zdh_sub_agencies:
        print(f'    · {sa["机构"]}: {sa["机构达人数"]}达人, GMV ¥{sa["直播GMV"]:,.2f}')
    print(f'  - 趋势窗口: {trend_dates[0]} ~ {trend_dates[-1]} ({len(trend_dates)} 天)')

    # 自动构建独立看板页面
    script_dir = os.path.dirname(os.path.abspath(__file__))
    build_script = os.path.join(script_dir, 'build_zidahao_standalone.py')
    if os.path.exists(build_script):
        print('')
        sp.run(['python3', build_script], cwd=script_dir)


if __name__ == '__main__':
    run()
